"""
Ledger reporting + export service mixin.

Two backing sources, unified into a single listing for the admin UI:

  1. `FinanceLedger`  — append-only record of *confirmed* financial events.
                        Source of truth for revenue / reports / exports.
                        Holds both admin-recorded payments (Cash / Manual
                        UPI from the office) AND parent UPI submissions
                        approved through the verification workflow.
  2. `Payment`        — every admin-recorded payment attempt. Used as a
                        fallback so PENDING / CANCELLED / FAILED states
                        without a ledger row still appear in the listing.

The list endpoint unions both so the admin sees every transaction state;
summary aggregation buckets them so revenue-only totals stay correct.
"""
from __future__ import annotations

import csv
import io
from datetime import datetime, date, time
from typing import Optional, Tuple, List, Dict, Any

from sqlalchemy import select, func, or_, exists
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.logger import logger  # noqa: F401
from app.models.finance import FinanceLedger, LedgerEntryType, Payment
from app.models.directory import Student
from app.models.academic import SchoolClass  # noqa: F401
from app.schemas.finance import LedgerSummary
from app.services.finance.ledger_helpers import (
    resolve_academic_year, backfill_missing_manual_ledger_entries,
)


def _academic_year_for(dt: Optional[datetime]) -> str:
    """Wrapper that gracefully handles `None` payment dates for orphan rows."""
    return resolve_academic_year((dt or datetime.utcnow()).date())


# Status buckets used by summary aggregation
COLLECTED_STATUS = "SUCCESS"
FAILED_STATUS = "FAILED"
REFUNDED_STATUS = "REFUNDED"
PENDING_STATUS = "PENDING"
CANCELLED_STATUS = "CANCELLED"

ALL_PAYMENT_STATUSES = (
    "SUCCESS",
    "PENDING",
    "FAILED",
    "CANCELLED",
    "REFUNDED",
    "PARTIALLY_REFUNDED",
)


class LedgerServiceMixin:
    def _build_ledger_query(
        self,
        institution_id: int,
        *,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
        student_id: Optional[int] = None,
        class_id: Optional[int] = None,
        fee_type: Optional[str] = None,
        payment_status: Optional[str] = None,
        payment_method: Optional[str] = None,
        academic_year: Optional[str] = None,
        min_amount: Optional[float] = None,
        max_amount: Optional[float] = None,
        search: Optional[str] = None,
    ):
        stmt = select(FinanceLedger).where(FinanceLedger.institution_id == institution_id)

        if date_from:
            stmt = stmt.where(FinanceLedger.payment_date >= date_from)
        if date_to:
            stmt = stmt.where(FinanceLedger.payment_date <= date_to)
        if student_id:
            stmt = stmt.where(FinanceLedger.student_id == student_id)
        if class_id:
            stmt = stmt.where(FinanceLedger.class_id == class_id)
        if fee_type:
            stmt = stmt.where(FinanceLedger.fee_type == fee_type)
        if payment_status:
            stmt = stmt.where(FinanceLedger.payment_status == payment_status)
        if payment_method:
            stmt = stmt.where(FinanceLedger.payment_method == payment_method)
        if academic_year:
            stmt = stmt.where(FinanceLedger.academic_year == academic_year)
        if min_amount is not None:
            stmt = stmt.where(FinanceLedger.amount >= min_amount)
        if max_amount is not None:
            stmt = stmt.where(FinanceLedger.amount <= max_amount)
        if search:
            like = f"%{search.strip()}%"
            stmt = stmt.where(
                or_(
                    FinanceLedger.student_name.ilike(like),
                    FinanceLedger.receipt_number.ilike(like),
                    FinanceLedger.external_reference.ilike(like),
                    FinanceLedger.class_name.ilike(like),
                )
            )

        return stmt

    async def get_ledger_summary(
        self,
        db: AsyncSession,
        institution_id: int,
        **filters,
    ) -> LedgerSummary:
        """
        Aggregate buckets for the summary cards. Revenue (`total_collected`)
        comes from FinanceLedger rows with status=SUCCESS, which is the
        unified source of truth — it covers admin-recorded payments AND
        manual UPI approvals from the verification workflow.

        Pending / cancelled buckets also include orphan Payment rows
        (admin-recorded attempts that never made it into the ledger).

        Runs a self-healing backfill first so any approved manual payment
        that didn't get mirrored (e.g. due to a transient failure during
        apply_decision) still shows up in revenue.
        """
        await backfill_missing_manual_ledger_entries(db, institution_id)

        base = self._build_ledger_query(institution_id, **filters)
        sub = base.subquery()

        agg_res = await db.execute(
            select(sub.c.payment_status, func.coalesce(func.sum(sub.c.amount), 0.0))
            .group_by(sub.c.payment_status)
        )
        amounts_by_status: Dict[str, float] = {
            row[0]: float(row[1] or 0.0) for row in agg_res.all()
        }

        count_res = await db.execute(select(func.count()).select_from(sub))
        ledger_count = int(count_res.scalar() or 0)

        orphan_totals, orphan_count = await self._orphan_payment_totals(
            db, institution_id, **filters
        )
        for status_key, total in orphan_totals.items():
            amounts_by_status[status_key] = amounts_by_status.get(status_key, 0.0) + total

        total_collected = amounts_by_status.get(COLLECTED_STATUS, 0.0)
        total_refunded = amounts_by_status.get(REFUNDED_STATUS, 0.0) + \
            amounts_by_status.get("PARTIALLY_REFUNDED", 0.0)

        return LedgerSummary(
            total_collected=total_collected,
            total_pending=amounts_by_status.get(PENDING_STATUS, 0.0),
            total_failed=amounts_by_status.get(FAILED_STATUS, 0.0),
            total_refunded=total_refunded,
            total_cancelled=amounts_by_status.get(CANCELLED_STATUS, 0.0),
            net_revenue=max(0.0, total_collected - total_refunded),
            transaction_count=ledger_count + orphan_count,
        )

    async def list_ledger_entries(
        self,
        db: AsyncSession,
        institution_id: int,
        *,
        skip: int = 0,
        limit: int = 50,
        **filters,
    ) -> Tuple[List[Dict[str, Any]], int]:
        # Self-heal: ensure every approved manual payment has its mirror
        # row before we read. Idempotent — does nothing on a healthy DB.
        await backfill_missing_manual_ledger_entries(db, institution_id)

        ledger_stmt = self._build_ledger_query(institution_id, **filters)
        ledger_res = await db.execute(ledger_stmt)
        ledger_rows: List[FinanceLedger] = list(ledger_res.scalars().all())

        orphan_rows = await self._fetch_orphan_payments(db, institution_id, **filters)

        payment_ids = [r.payment_id for r in ledger_rows if r.payment_id]
        refund_index = await self._build_refund_index(db, institution_id, payment_ids)

        entries: List[Dict[str, Any]] = [
            self._ledger_row_to_dict(r, refund_index) for r in ledger_rows
        ]
        entries.extend(self._payment_to_ledger_dict(*row) for row in orphan_rows)

        entries.sort(key=lambda e: e["payment_date"] or datetime.min, reverse=True)
        total = len(entries)
        page = entries[skip : skip + limit]
        return page, total

    # --- Orphan-payment helpers ------------------------------------------------

    def _apply_orphan_filters(
        self,
        stmt,
        *,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
        student_id: Optional[int] = None,
        class_id: Optional[int] = None,
        fee_type: Optional[str] = None,
        payment_status: Optional[str] = None,
        payment_method: Optional[str] = None,
        academic_year: Optional[str] = None,
        min_amount: Optional[float] = None,
        max_amount: Optional[float] = None,
        search: Optional[str] = None,
    ):
        if date_from:
            stmt = stmt.where(Payment.created_at >= date_from)
        if date_to:
            stmt = stmt.where(Payment.created_at <= date_to)
        if student_id:
            stmt = stmt.where(Payment.student_id == student_id)
        if class_id:
            stmt = stmt.where(Student.school_class_id == class_id)
        if payment_status:
            stmt = stmt.where(Payment.status == payment_status)
        if payment_method:
            stmt = stmt.where(Payment.payment_mode == payment_method)
        if min_amount is not None:
            stmt = stmt.where(Payment.amount >= min_amount)
        if max_amount is not None:
            stmt = stmt.where(Payment.amount <= max_amount)
        if search:
            like = f"%{search.strip()}%"
            stmt = stmt.where(Student.name.ilike(like))
        # `fee_type` and `academic_year` don't apply to orphan payments
        if fee_type:
            stmt = stmt.where(False)
        if academic_year:
            stmt = stmt.where(False)
        return stmt

    async def _fetch_orphan_payments(
        self,
        db: AsyncSession,
        institution_id: int,
        **filters,
    ) -> List[Tuple[Payment, Optional[str], Optional[int]]]:
        no_ledger_clause = ~exists().where(FinanceLedger.payment_id == Payment.id)
        stmt = (
            select(Payment, Student.name, Student.school_class_id)
            .join(Student, Payment.student_id == Student.id)
            .where(Payment.institution_id == institution_id)
            .where(no_ledger_clause)
        )
        stmt = self._apply_orphan_filters(stmt, **filters)
        res = await db.execute(stmt)
        return [(row[0], row[1], row[2]) for row in res.all()]

    async def _orphan_payment_totals(
        self,
        db: AsyncSession,
        institution_id: int,
        **filters,
    ) -> Tuple[Dict[str, float], int]:
        no_ledger_clause = ~exists().where(FinanceLedger.payment_id == Payment.id)
        stmt = (
            select(Payment.status, func.coalesce(func.sum(Payment.amount), 0.0), func.count())
            .join(Student, Payment.student_id == Student.id)
            .where(Payment.institution_id == institution_id)
            .where(no_ledger_clause)
            .group_by(Payment.status)
        )
        stmt = self._apply_orphan_filters(stmt, **filters)
        res = await db.execute(stmt)
        totals: Dict[str, float] = {}
        count = 0
        for status_key, amt, n in res.all():
            if status_key:
                totals[status_key] = float(amt or 0.0)
                count += int(n or 0)
        return totals, count

    async def _build_refund_index(
        self,
        db: AsyncSession,
        institution_id: int,
        payment_ids: List[int],
    ) -> Dict[int, float]:
        if not payment_ids:
            return {}
        res = await db.execute(
            select(
                FinanceLedger.payment_id,
                func.coalesce(func.sum(FinanceLedger.amount), 0.0),
            )
            .where(FinanceLedger.institution_id == institution_id)
            .where(FinanceLedger.entry_type == LedgerEntryType.REFUND)
            .where(FinanceLedger.payment_id.in_(payment_ids))
            .group_by(FinanceLedger.payment_id)
        )
        return {row[0]: float(row[1] or 0.0) for row in res.all()}

    # --- Row → dict shapers ----------------------------------------------------

    @staticmethod
    def _refund_status(amount: float, refunded: float) -> Tuple[Optional[str], float]:
        if refunded <= 0.0001:
            return None, 0.0
        if refunded + 0.0001 >= amount:
            return "REFUNDED", refunded
        return "PARTIALLY_REFUNDED", refunded

    @staticmethod
    def _ledger_row_to_dict(
        row: FinanceLedger,
        refund_index: Dict[int, float],
    ) -> Dict[str, Any]:
        entry_type = row.entry_type.value if hasattr(row.entry_type, "value") else row.entry_type
        amount = float(row.amount or 0.0)
        net = float(row.net_amount if row.net_amount is not None else amount)

        refund_status: Optional[str] = None
        refunded_amount = 0.0
        if entry_type == "PAYMENT" and row.payment_id and row.payment_id in refund_index:
            refund_status, refunded_amount = LedgerServiceMixin._refund_status(
                amount, refund_index[row.payment_id]
            )

        error_message: Optional[str] = None
        if row.payment_status in {"FAILED", "CANCELLED"}:
            error_message = row.notes

        has_receipt = (
            entry_type == "PAYMENT"
            and row.payment_status == "SUCCESS"
            and row.id is not None
            and row.id > 0
        )

        return {
            "id": row.id,
            "receipt_number": row.receipt_number,
            "entry_type": entry_type,
            "payment_id": row.payment_id,
            "manual_payment_request_id": row.manual_payment_request_id,
            "student_id": row.student_id,
            "student_name": row.student_name,
            "class_id": row.class_id,
            "class_name": row.class_name,
            "fee_type": row.fee_type,
            "academic_year": row.academic_year,
            "amount": amount,
            "gateway_fee": float(row.gateway_fee or 0.0),
            "net_amount": net,
            "payment_method": row.payment_method,
            "payment_status": row.payment_status,
            "payment_date": row.payment_date,
            "notes": row.notes,
            "transaction_id": row.external_reference,
            "refund_status": refund_status,
            "refunded_amount": refunded_amount or None,
            "error_message": error_message,
            "has_receipt": has_receipt,
        }

    @staticmethod
    def _payment_to_ledger_dict(
        payment: Payment,
        student_name: Optional[str],
        student_class_id: Optional[int],
    ) -> Dict[str, Any]:
        """
        Synthesise a ledger-shaped row from an orphan Payment so PENDING /
        CANCELLED / FAILED admin-recorded attempts show up in the listing.
        """
        amount = float(payment.amount or 0.0)
        status_val = payment.status or "PENDING"
        return {
            "id": -int(payment.id),
            "receipt_number": f"PEND-{payment.id:06d}",
            "entry_type": "PAYMENT",
            "payment_id": payment.id,
            "manual_payment_request_id": None,
            "student_id": payment.student_id,
            "student_name": student_name or f"Student #{payment.student_id}",
            "class_id": student_class_id,
            "class_name": None,
            "fee_type": "TUITION",
            "academic_year": _academic_year_for(payment.created_at),
            "amount": amount,
            "gateway_fee": 0.0,
            "net_amount": amount,
            "payment_method": payment.payment_mode or "CASH",
            "payment_status": status_val,
            "payment_date": payment.created_at,
            "notes": payment.note,
            "transaction_id": None,
            "refund_status": None,
            "refunded_amount": None,
            "error_message": payment.note if status_val in {"FAILED", "CANCELLED"} else None,
            "has_receipt": False,
        }

    async def fetch_unified_for_export(
        self,
        db: AsyncSession,
        institution_id: int,
        *,
        date_from: datetime,
        date_to: datetime,
        **filters,
    ) -> List[Dict[str, Any]]:
        all_filters = dict(filters)
        all_filters["date_from"] = date_from
        all_filters["date_to"] = date_to
        entries, _ = await self.list_ledger_entries(
            db, institution_id, skip=0, limit=10_000, **all_filters
        )
        entries.sort(key=lambda e: e["payment_date"] or datetime.min)
        return entries

    async def fetch_ledger_for_export(
        self,
        db: AsyncSession,
        institution_id: int,
        *,
        date_from: datetime,
        date_to: datetime,
        **filters,
    ) -> List[FinanceLedger]:
        """Return every row in range (no pagination) for export use."""
        stmt = self._build_ledger_query(
            institution_id, date_from=date_from, date_to=date_to, **filters
        )
        stmt = stmt.order_by(FinanceLedger.payment_date.asc())
        res = await db.execute(stmt)
        return res.scalars().all()

    async def get_ledger_facets(
        self,
        db: AsyncSession,
        institution_id: int,
    ) -> dict:
        st_res = await db.execute(
            select(FinanceLedger.payment_status)
            .where(FinanceLedger.institution_id == institution_id)
            .distinct()
        )
        ledger_statuses = {row[0] for row in st_res.all() if row[0]}
        pay_res = await db.execute(
            select(Payment.status)
            .where(Payment.institution_id == institution_id)
            .distinct()
        )
        payment_statuses = {row[0] for row in pay_res.all() if row[0]}
        statuses = sorted(ledger_statuses | payment_statuses | set(ALL_PAYMENT_STATUSES))

        m_res = await db.execute(
            select(FinanceLedger.payment_method)
            .where(FinanceLedger.institution_id == institution_id)
            .distinct()
        )
        methods = sorted([row[0] for row in m_res.all() if row[0]])

        ft_res = await db.execute(
            select(FinanceLedger.fee_type)
            .where(FinanceLedger.institution_id == institution_id)
            .distinct()
        )
        fee_types = sorted([row[0] for row in ft_res.all() if row[0]])

        ay_res = await db.execute(
            select(FinanceLedger.academic_year)
            .where(FinanceLedger.institution_id == institution_id)
            .distinct()
        )
        academic_years = sorted(
            [row[0] for row in ay_res.all() if row[0]], reverse=True
        )

        from app.models.academic import SchoolClass
        cls_res = await db.execute(
            select(
                SchoolClass.id,
                SchoolClass.display_name,
                SchoolClass.grade_id,
                SchoolClass.section_id,
            )
            .where(SchoolClass.institution_id == institution_id)
            .order_by(SchoolClass.display_name)
        )
        classes = [
            {
                "id": row[0],
                "display_name": row[1] or f"Class {row[2]}",
                "grade_id": row[2],
                "section_id": row[3],
            }
            for row in cls_res.all()
        ]

        first_res = await db.execute(
            select(func.min(FinanceLedger.payment_date)).where(
                FinanceLedger.institution_id == institution_id
            )
        )
        last_res = await db.execute(
            select(func.max(FinanceLedger.payment_date)).where(
                FinanceLedger.institution_id == institution_id
            )
        )
        earliest = first_res.scalar()
        latest = last_res.scalar()

        return {
            "statuses": statuses,
            "methods": methods,
            "fee_types": fee_types,
            "academic_years": academic_years,
            "classes": classes,
            "earliest_payment_date": earliest.isoformat() if earliest else None,
            "latest_payment_date": latest.isoformat() if latest else None,
        }


# --- Date-range helpers ---

def normalise_date_range(
    start: date, end: date
) -> Tuple[datetime, datetime]:
    if end < start:
        raise ValueError("date_to must be >= date_from")
    start_dt = datetime.combine(start, time.min)
    end_dt = datetime.combine(end, time.max)
    return start_dt, end_dt


# --- Exporters: CSV, Excel, PDF ---

def _row_dict(entry: FinanceLedger) -> dict:
    return {
        "Receipt #": entry.receipt_number,
        "Date": entry.payment_date.strftime("%Y-%m-%d %H:%M") if entry.payment_date else "",
        "Student": entry.student_name,
        "Student ID": entry.student_id,
        "Class": entry.class_name or "",
        "Fee Type": entry.fee_type or "",
        "Academic Year": entry.academic_year,
        "Amount (INR)": entry.amount,
        "Gateway Fee": entry.gateway_fee or 0.0,
        "Net Amount": entry.net_amount if entry.net_amount is not None else entry.amount,
        "Method": entry.payment_method,
        "Status": entry.payment_status,
        "Reference": entry.external_reference or "",
        "Notes": entry.notes or "",
    }


def export_csv(entries: List[FinanceLedger], date_from: date, date_to: date) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Finance Ledger Export"])
    writer.writerow([f"Range: {date_from.isoformat()} to {date_to.isoformat()}"])
    writer.writerow([f"Generated: {datetime.utcnow().isoformat()}Z"])
    writer.writerow([])

    if not entries:
        writer.writerow(["No payments in this date range."])
        return buf.getvalue().encode("utf-8")

    headers = list(_row_dict(entries[0]).keys())
    writer.writerow(headers)
    total_amount = 0.0
    total_net = 0.0
    for e in entries:
        row = _row_dict(e)
        writer.writerow(row.values())
        total_amount += float(row["Amount (INR)"] or 0)
        total_net += float(row["Net Amount"] or 0)

    writer.writerow([])
    totals_row = ["TOTALS"] + [""] * (len(headers) - 1)
    totals_row[headers.index("Amount (INR)")] = total_amount
    totals_row[headers.index("Net Amount")] = total_net
    writer.writerow(totals_row)
    return buf.getvalue().encode("utf-8")


def export_excel(entries: List[FinanceLedger], date_from: date, date_to: date) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError as e:
        raise RuntimeError(
            "openpyxl is required for Excel exports. Install with: pip install openpyxl"
        ) from e

    wb = Workbook()
    ws = wb.active
    ws.title = "Finance Ledger"

    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    total_font = Font(bold=True)

    ws.append(["Finance Ledger Export"])
    ws["A1"].font = title_font
    ws.append([f"Range: {date_from.isoformat()} → {date_to.isoformat()}"])
    ws.append([f"Generated: {datetime.utcnow().isoformat()}Z"])
    ws.append([])

    if not entries:
        ws.append(["No payments in this date range."])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    headers = list(_row_dict(entries[0]).keys())
    ws.append(headers)
    header_row_idx = ws.max_row
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    total_amount = 0.0
    total_net = 0.0
    for e in entries:
        row = _row_dict(e)
        ws.append(list(row.values()))
        total_amount += float(row["Amount (INR)"] or 0)
        total_net += float(row["Net Amount"] or 0)

    ws.append([])
    totals_row = ["TOTALS"] + [""] * (len(headers) - 1)
    amount_col = headers.index("Amount (INR)")
    net_col = headers.index("Net Amount")
    totals_row[amount_col] = total_amount
    totals_row[net_col] = total_net
    ws.append(totals_row)
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=ws.max_row, column=col_idx).font = total_font

    for col_cells in ws.columns:
        col_letter = col_cells[0].column_letter
        max_len = max(
            (len(str(c.value)) for c in col_cells if c.value is not None),
            default=10,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_pdf(entries: List[FinanceLedger], date_from: date, date_to: date) -> bytes:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        )
    except ImportError as e:
        raise RuntimeError(
            "reportlab is required for PDF exports."
        ) from e

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=20,
        rightMargin=20,
        topMargin=20,
        bottomMargin=20,
    )
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("<b>Finance Ledger Export</b>", styles["Title"]))
    story.append(Paragraph(
        f"Range: {date_from.isoformat()} &nbsp;&rarr;&nbsp; {date_to.isoformat()}",
        styles["Normal"],
    ))
    story.append(Paragraph(
        f"Generated: {datetime.utcnow().isoformat()}Z",
        styles["Normal"],
    ))
    story.append(Spacer(1, 12))

    if not entries:
        story.append(Paragraph("No payments in this date range.", styles["Normal"]))
        doc.build(story)
        return buf.getvalue()

    headers = [
        "Receipt #",
        "Date",
        "Student",
        "Class",
        "Method",
        "Status",
        "Amount",
        "Net",
    ]
    table_data = [headers]
    total_amount = 0.0
    total_net = 0.0
    for e in entries:
        amount = float(e.amount or 0.0)
        net = float(e.net_amount if e.net_amount is not None else amount)
        total_amount += amount
        total_net += net
        table_data.append([
            e.receipt_number,
            e.payment_date.strftime("%Y-%m-%d %H:%M") if e.payment_date else "",
            e.student_name,
            e.class_name or "",
            e.payment_method,
            e.payment_status,
            f"₹{amount:,.2f}",
            f"₹{net:,.2f}",
        ])
    table_data.append([
        "", "", "", "", "", "TOTAL",
        f"₹{total_amount:,.2f}", f"₹{total_net:,.2f}",
    ])

    tbl = Table(table_data, repeatRows=1, hAlign="LEFT")
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (-2, 0), (-1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F3F4F6")),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
    ]))
    story.append(tbl)

    doc.build(story)
    return buf.getvalue()
